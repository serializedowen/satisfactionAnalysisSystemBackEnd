from openpyxl import load_workbook#读取excel文件
import os

import json
# file 文件
# file_display_name 文件展示名
# score_obj 选项和分值的对应关系

# title_range 表头范围
# choose_range 可选项的范围


# y_range 内生指标范围
# x_range 外生指标范围


def procession(file_path, score_obj, titleRangeStart, titleRangeEnd, chooseRangeStart, chooseRangeEnd, xRangeStart, xRangeEnd, yRangeStart, yRangeEnd):
    try:
        file_path = file_path
        wb2 = load_workbook(file_path)
        ws = wb2.get_sheet_by_name(wb2.sheetnames[0])

        # 选项和分值的对应关系
        score_obj = json.loads(score_obj)
        max_score = max(score_obj.values())
        min_score = min(score_obj.values())

        # 内生潜变量范围
        y_data = []
        cell_ydata = ws[yRangeStart:yRangeEnd] # y_range
        for y_item in cell_ydata:
            for i in y_item:
                y_data.append(i.value)

        titles=[]
        cell_titles = ws[titleRangeStart:titleRangeEnd] # title_range
        for c in cell_titles:
            for i in c:
                if i.value:
                    titles.append(i.value)


        # 这里需要输入表头选项
        cell_choose = ws[chooseRangeStart:chooseRangeEnd] # choose_range
        # 表头选项的数组
        choose=[]
        for cell in cell_choose:
            for i in cell:
                choose.append(i.value[0])

        # 这里需要输入表格的内容
        cell_range2 = ws[xRangeStart:xRangeEnd] # x_range
        # 表格数据矩阵
        data=[]
        for cell2 in range(len(cell_range2)):
            row = []
            for i2 in range(len(cell_range2[cell2])):
                if cell_range2[cell2][i2].value:
                    # 记性 max-min 数据处理，然后保留6位小数
                    item = format((score_obj[choose[i2]] - min_score) / (max_score - min_score), '.6f')
                    row.append(item)
            data.append(row)


        res_str = ''
        for i in range(len(data)):
            row_str = '     '.join(data[i])

            # 如果用户不填值，这条数据就不算数
            if y_data[i]:
                if i == len(data)-1:
                    res_str = res_str + row_str + '     ' + format((y_data[i] - (len(titles) * min_score)) / ((len(titles) * max_score) - (len(titles) * min_score)), '.6f')
                else:
                    res_str = res_str + row_str + '     ' + format((y_data[i] - (len(titles) * min_score)) / ((len(titles) * max_score) - (len(titles) * min_score)), '.6f') + '\n'

        for k in range(len(data)):
            if y_data[k]:
                item = format(
                    (y_data[k] - (len(titles) * min_score)) / ((len(titles) * max_score) - (len(titles) * min_score)),
                    '.6f')
                data[k].append(item)

        outfile_name = file_path + '.dat'
        f = open(outfile_name, 'w+')
        f.write(res_str)
        f.close()
        return {"titleList": titles, "xyRange": data}
    except:
        return False
